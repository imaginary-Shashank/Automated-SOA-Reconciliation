"""Excel operations for reading summary reports and writing to SOA files."""

import logging
from pathlib import Path
import openpyxl
import xlwings as xw

from schemas import TerminationCharge, ReconciliationBatch

logger = logging.getLogger(__name__)

def read_summary_report(file_path: str | Path) -> ReconciliationBatch:
    """
    Reads the termination summary Excel file and yields a validated batch of charges.
    """
    path = Path(file_path)
    if not path.is_file():
        raise FileNotFoundError(f"Summary report not found at {path}")

    logger.info(f"Loading summary report: {path.name}")
    
    # data_only=True ensures we extract the raw values, not Excel formulas
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active 
    
    records = []
    
    # max_col=4 captures Client, Bill, Status, and Comment
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        client_val, bill_val, status_val, comment_val = row
        
        # Stop processing if we hit an entirely blank row at the bottom of the data
        if client_val is None and bill_val is None:
            break
            
        try:
            charge = TerminationCharge(
                carrier_name=str(client_val) if client_val else "",
                amount=bill_val,
                status=str(status_val) if status_val else None,
                comment=str(comment_val) if comment_val else None,
                source_row=row_idx
            )
            records.append(charge)
        except Exception as e:
            logger.error(f"Row {row_idx} failed validation: {client_val} | Error: {e}")
            
    logger.info(f"Successfully extracted {len(records)} clean records.")
    return ReconciliationBatch(records=records)

def update_soa_file(soa_path: Path, charge: TerminationCharge, start_date: str, end_date: str):
    """
    Safely injects a new billing record into the target SOA file.
    Preserves existing formulas by not using data_only=True.
    """
    logger.info(f"Opening {soa_path.name} for write operations...")
    
    # Load without data_only to protect existing formulas in the sheet
    wb = openpyxl.load_workbook(soa_path)
    sheet = wb.active
    
    # Data starts at row 11 according to the SOA structure (*Note: Specific to the firm, you can change accordingly)
    target_row = 11
    
    # *Scan Column F (Index 6) downwards to find the first empty "Billing Start Date"
    while True:
        cell_val = sheet.cell(row=target_row, column=6).value
        # Break if the cell is completely empty or just whitespace
        if cell_val is None or str(cell_val).strip() == "":
            break
        target_row += 1
        
    logger.info(f"Targeting Row {target_row} for new entry.")
    
    # *1. Input the billing dates (Columns F and G)
    sheet.cell(row=target_row, column=6).value = start_date
    sheet.cell(row=target_row, column=7).value = end_date
    
    # *2. Input the charges (Column K). 
    # Cast Decimal to float so Excel natively recognizes it as a number, not text.
    sheet.cell(row=target_row, column=11).value = float(charge.amount)
    
    # *3. Write *unbilled at the end of the table (Column N)
    sheet.cell(row=target_row, column=14).value = "*Unbilled"
    
    # Save the file
    wb.save(soa_path)
    logger.info(f" Successfully wrote ${charge.amount} to {soa_path.name}")

def update_summary_report(report_path: Path | str, updates: list[tuple[int, str, int]]):
    """
    Writes the final processing statuses back to a new summary report.
    Uses 0 and 1 logic to determine output, and hardcodes the company 
    names to prevent formula corruption.
    """
    path = Path(report_path)
    logger.info(f"Writing final statuses back to summary report: {path.name}...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    
    # Apply all updates using the 0 and 1 logic
    for row_idx, company_name, amount, status_code in updates:
        # 1. Hardcode the company name and amount to prevent them from vanishing
        sheet.cell(row=row_idx, column=1).value = company_name
        sheet.cell(row=row_idx, column=2).value = float(amount)
        
        # 2. Apply the logic based on the 0 or 1 status code
        if status_code == 1:
            sheet.cell(row=row_idx, column=3).value = "Done"
            sheet.cell(row=row_idx, column=4).value = "" # Clear comment
        elif status_code == 0:
            sheet.cell(row=row_idx, column=4).value = "SOA not found"
            
    # Save as a NEW file so the original source remains untouched
    new_path = path.parent / f"{path.stem}_PROCESSED{path.suffix}"
    wb.save(new_path)
    logger.info(f" Processed report saved safely as: {new_path.name}")
