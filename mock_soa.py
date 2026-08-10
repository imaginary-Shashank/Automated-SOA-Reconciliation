"""Generates varied, realistic mock SOA files for local pipeline testing."""

import shutil
import random
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl

def generate_varied_mocks():
    raw_names = [
        "Cornwall Telecom Limited - YO0619",
        "Data Access Solutions Inc - YO0619",
        "DIAL TEL - NA0419",
        "EvoiceBD Ltd - YO0619",
        "Evolve Tech - KP0125",
        "Figoria/Wingz Telecom - NA0419",
        "G5 Telekom - KP0125",
        "Galaxy Nexus AI Group - YO0619",
        "Globalcarrier telecom - AN1018",
        "Globe Teleservices - YO0619",
        "Globlink Technology Co Ltd - KP0125",
        "GM TELECOM CORP - YO0619",
        "Grupo Marpica SA - NA0419",
        "Hostertel - KP0125",
        "IGlobe Telecom Networks - YO0619",
        "Jaisotel Limited - DO0620"
    ]

    template_path = Path(" ")# Enter a file name here present in your directory for refrence of the mock SOAs
    output_dir = Path(" ")
    
    # Reset the directory for a clean slate
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir()
    
    if not template_path.exists():
        print(f" Error: Template '{template_path.name}' not found.")
        return

    print(f"Generating varied mock SOAs in ./{output_dir.name}/ ...\n")

    selected_names = random.sample(raw_names, 12)
    
    for record in selected_names:
        company_name = record.split(" - ")[0].strip()
        new_filename = f"{company_name} & Aiwo SOA.xlsx"
        new_filepath = output_dir / new_filename
        
        shutil.copy(template_path, new_filepath)
        
        try:
            wb = openpyxl.load_workbook(new_filepath)
            sheet = wb.active
            
            # 1. Update Identity
            sheet['C6'] = company_name
            email_prefix = company_name.split()[0].lower().replace("/", "")
            sheet['C7'] = f"billing@{email_prefix}.com"
            
            # 2. Safely clear old data (Rows 11 to 30) avoiding MergedCells
            data_columns = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14] 
            for row in range(11, 31):
                for col in data_columns:
                    cell = sheet.cell(row=row, column=col)
                    # SAFEGUARD: Only write to the cell if it's NOT a MergedCell
                    if type(cell).__name__ != 'MergedCell':
                        cell.value = None
            
            # 3. Generate random historical data (between 1 and 6 weeks of history)
            num_history_rows = random.randint(1, 6)
            base_date = datetime(2026, 4, 13) 
            
            for i in range(num_history_rows):
                current_row = 11 + i
                
                # Create weekly billing dates
                start_date = base_date + timedelta(days=7*i)
                end_date = start_date + timedelta(days=6)
                start_str = start_date.strftime("%d-%b-%Y")
                end_str = end_date.strftime("%d-%b-%Y")
                
                # Generate realistic random amounts
                billed_amt = round(random.uniform(10.0, 200.0), 4)
                charges_amt = round(billed_amt + random.uniform(-5.0, 5.0), 4)
                
                # Helper function for safe writing
                def safe_write(r, c, val):
                    target_cell = sheet.cell(row=r, column=c)
                    if type(target_cell).__name__ != 'MergedCell':
                        target_cell.value = val

                # Inject Partner Data (Columns A-E)
                safe_write(current_row, 1, start_str)
                safe_write(current_row, 2, end_str)
                safe_write(current_row, 3, billed_amt)
                safe_write(current_row, 4, 0) # Payment Received
                safe_write(current_row, 5, None) # Payment Date
                
                # Inject Our Data (Columns F-K)
                safe_write(current_row, 6, start_str)
                safe_write(current_row, 7, end_str)
                safe_write(current_row, 8, billed_amt)
                safe_write(current_row, 9, 0) # Payment Sent
                safe_write(current_row, 10, None) # Payment Date
                safe_write(current_row, 11, charges_amt)
                
                # Randomly mark older invoices as *unbilled
                if random.choice([True, False]):
                    safe_write(current_row, 14, "*unbilled")
                    
            wb.save(new_filepath)
            print(f" Generated: {new_filename} ({num_history_rows} past records)")
            
        except Exception as e:
            print(f" Failed to update {new_filename}: {e}")

    print("\n Varied mock generation complete!")

if __name__ == "__main__":
    generate_varied_mocks()
